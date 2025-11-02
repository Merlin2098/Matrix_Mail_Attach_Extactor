"""
Extractor de adjuntos de correos de Outlook.

FASE 2 - Hereda de BackendBase
"""

import os
import sys
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import List, Optional

import pythoncom
import win32com.client
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from backend_base import (
    BackendBase,
    FaseProceso as FaseBase,
    NivelMensaje,
    EstadoProceso,
    EstadisticasBase
)


# ========================================
# RE-EXPORTAR PARA USO EXTERNO
# ========================================
__all__ = [
    'ExtractorAdjuntosOutlook',
    'FaseProceso',
    'NivelMensaje',
    'EstadoProceso',  # Re-exportado desde backend_base
    'EstadisticasExtraccion'
]


# ========================================
# ENUMS ESPECÍFICOS
# ========================================

class FaseProceso(Enum):
    """Fases específicas del proceso de extracción"""
    INICIAL = "inicial"
    FILTRADO = "filtrado"
    DESCARGA = "descarga"
    FINALIZACION = "finalizacion"


# ========================================
# ESTADÍSTICAS ESPECÍFICAS
# ========================================

@dataclass
class EstadisticasExtraccion(EstadisticasBase):
    """Estadísticas del proceso de extracción"""
    total_correos: int = 0
    correos_procesados: int = 0
    adjuntos_descargados: int = 0
    adjuntos_fallidos: int = 0
    tamaño_total_mb: float = 0.0
    archivos_descargados: List[dict] = field(default_factory=list)
    
    @property
    def tasa_exito(self) -> float:
        """Porcentaje de éxito"""
        total = self.adjuntos_descargados + self.adjuntos_fallidos
        return (self.adjuntos_descargados / total * 100) if total > 0 else 0.0


# ========================================
# CLASE PRINCIPAL
# ========================================

class ExtractorAdjuntosOutlook(BackendBase):
    """
    Extractor de adjuntos de correos de Outlook con filtrado avanzado.
    
    FASE 2 - Hereda de BackendBase:
    - Reutiliza callbacks, control de flujo, utilidades
    - Solo implementa lógica específica de Outlook
    """
    
    def __init__(self, 
                 callback_mensaje=None,
                 callback_progreso=None, 
                 callback_estado=None):
        """Inicializa el extractor"""
        super().__init__(callback_mensaje, callback_progreso, callback_estado)
        
        self.estadisticas = EstadisticasExtraccion()
        
        # Configuración específica
        self.config = {
            "min_lote": 10,
            "max_lote": 50,
            "max_reintentos": 3,
            "pausa_entre_lotes": 0.5,
            "liberar_memoria_cada": 5
        }
    
    # ========================================
    # IMPLEMENTACIÓN DE MÉTODOS ABSTRACTOS
    # ========================================
    
    def validar_parametros(self, frases: List[str], destino: str, 
                          outlook_folder: str, fecha_inicio: datetime, 
                          fecha_fin: datetime) -> tuple[bool, str]:
        """Valida los parámetros de extracción"""
        
        # Validar carpeta destino
        if not destino or not destino.strip():
            return False, "Debe seleccionar una carpeta de destino"
        
        # Validar carpeta Outlook
        if not outlook_folder or not outlook_folder.strip():
            return False, "Debe seleccionar una bandeja de correo"
        
        # Validar fechas
        es_valido, mensaje = self._validar_rango_fechas(fecha_inicio, fecha_fin)
        if not es_valido:
            return False, mensaje
        
        # Crear carpeta destino si no existe
        try:
            Path(destino).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            return False, f"No se puede crear la carpeta de destino: {str(e)}"
        
        # Aviso si no hay frases
        if not frases:
            self._enviar_mensaje(
                FaseProceso.INICIAL,
                NivelMensaje.WARNING,
                "No se especificaron frases: se descargarán todos los adjuntos"
            )
        
        return True, ""
    
    def _procesar_principal(self, frases: List[str], destino: str,
                           outlook_folder: str, fecha_inicio: datetime,
                           fecha_fin: datetime) -> dict:
        """Procesamiento principal de extracción"""
        
        # Resetear estadísticas
        self.estadisticas = EstadisticasExtraccion()
        self.estadisticas.tiempo_inicio = datetime.now()
        
        try:
            # Crear log
            self._crear_log_archivo(destino, "log_extraccion")
            
            # Conectar a Outlook
            namespace = self._conectar_outlook()
            carpeta = self._obtener_carpeta(namespace, outlook_folder)
            
            # Filtrar correos
            correos_filtrados = self._filtrar_correos(
                carpeta, frases, fecha_inicio, fecha_fin
            )
            
            if not correos_filtrados:
                self._enviar_mensaje(
                    FaseProceso.FILTRADO,
                    NivelMensaje.WARNING,
                    "No se encontraron correos con los criterios especificados"
                )
                self.estadisticas.tiempo_fin = datetime.now()
                return self._generar_reporte()
            
            # Descargar adjuntos
            self._descargar_adjuntos(correos_filtrados, destino)
            
            # Finalizar
            self.estadisticas.tiempo_fin = datetime.now()
            
            # Generar reportes
            self._finalizar_log_archivo(self._generar_reporte())
            self._generar_excel_listado(destino)
            
            self._enviar_mensaje(
                FaseProceso.FINALIZACION,
                NivelMensaje.SUCCESS,
                f"Proceso completado: {self.estadisticas.adjuntos_descargados} adjuntos descargados"
            )
            
            return self._generar_reporte()
            
        except InterruptedError:
            self.estadisticas.tiempo_fin = datetime.now()
            self._finalizar_log_archivo(self._generar_reporte())
            raise
            
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    
    def _generar_reporte(self) -> dict:
        """Genera reporte final de estadísticas"""
        return {
            "total_correos": self.estadisticas.total_correos,
            "correos_procesados": self.estadisticas.correos_procesados,
            "adjuntos_descargados": self.estadisticas.adjuntos_descargados,
            "adjuntos_fallidos": self.estadisticas.adjuntos_fallidos,
            "tamaño_total_mb": self.estadisticas.tamaño_total_mb,
            "tiempo_total": self.estadisticas.tiempo_total,
            "tasa_exito": self.estadisticas.tasa_exito
        }
    
    # ========================================
    # MÉTODOS ESPECÍFICOS DE OUTLOOK
    # ========================================
    
    def _conectar_outlook(self):
        """Establece conexión con Outlook"""
        try:
            pythoncom.CoInitialize()
            outlook = win32com.client.Dispatch("Outlook.Application")
            namespace = outlook.GetNamespace("MAPI")
            self._enviar_mensaje(
                FaseProceso.INICIAL,
                NivelMensaje.SUCCESS,
                "Conectado a Outlook exitosamente"
            )
            return namespace
        except Exception as e:
            error_msg = f"Error al conectar con Outlook: {str(e)}"
            self._enviar_mensaje(
                FaseProceso.INICIAL,
                NivelMensaje.ERROR,
                error_msg
            )
            raise ConnectionError(error_msg)
    
    def _obtener_carpeta(self, namespace, ruta_carpeta: str):
        """Obtiene una carpeta de Outlook por su ruta"""
        try:
            self._enviar_mensaje(
                FaseProceso.INICIAL,
                NivelMensaje.INFO,
                f"Buscando carpeta: {ruta_carpeta}"
            )
            
            partes = ruta_carpeta.split("\\")
            carpeta = namespace.Folders[partes[0]]
            
            for parte in partes[1:]:
                carpeta = carpeta.Folders[parte]
            
            self._enviar_mensaje(
                FaseProceso.INICIAL,
                NivelMensaje.SUCCESS,
                f"Carpeta encontrada: {carpeta.Name} ({carpeta.Items.Count} elementos)"
            )
            
            return carpeta
        except Exception as e:
            error_msg = f"Error al acceder a la carpeta '{ruta_carpeta}': {str(e)}"
            self._enviar_mensaje(
                FaseProceso.INICIAL,
                NivelMensaje.ERROR,
                error_msg
            )
            raise ValueError(error_msg)
    
    def _filtrar_correos(self, carpeta, frases: List[str], 
                        fecha_inicio: datetime, fecha_fin: datetime) -> List:
        """Filtra correos según criterios"""
        
        self.fase_actual = FaseProceso.FILTRADO
        self._cambiar_estado(EstadoProceso.FILTRANDO)  # ← Cambiado
        
        self._enviar_mensaje(
            FaseProceso.FILTRADO,
            NivelMensaje.INFO,
            f"Iniciando filtrado de correos (Total: {carpeta.Items.Count})"
        )
        
        # ✅ CORRECCIÓN: Normalizar fechas para comparar solo fecha (sin hora)
        # Fecha inicio: 00:00:00 del día
        fecha_inicio_normalizada = fecha_inicio.replace(hour=0, minute=0, second=0, microsecond=0)
        # Fecha fin: 23:59:59 del día
        fecha_fin_normalizada = fecha_fin.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        self._enviar_mensaje(
            FaseProceso.FILTRADO,
            NivelMensaje.INFO,
            f"Rango de fechas solicitado: {fecha_inicio_normalizada.date()} a {fecha_fin_normalizada.date()}"
        )
        
        # ✅ NUEVO: Detectar primera y última fecha en la bandeja
        items = carpeta.Items
        items.Sort("[ReceivedTime]", True)  # Ordenar descendente (más reciente primero)
        
        total_items = items.Count
        primera_fecha = None
        ultima_fecha = None
        
        # Detectar rango real de fechas en la bandeja
        if total_items > 0:
            try:
                # Última fecha (más reciente) - primer item (índice 1 en win32com)
                for i in range(1, min(total_items + 1, 20)):  # Revisar primeros 20
                    try:
                        item = items.Item(i)
                        if hasattr(item, 'ReceivedTime'):
                            ultima_fecha = item.ReceivedTime.replace(tzinfo=None)
                            break
                    except:
                        continue
                
                # Primera fecha (más antigua) - últimos items
                for i in range(total_items, max(total_items - 20, 0), -1):  # Revisar últimos 20
                    try:
                        item = items.Item(i)
                        if hasattr(item, 'ReceivedTime'):
                            primera_fecha = item.ReceivedTime.replace(tzinfo=None)
                            break
                    except:
                        continue
                
                if primera_fecha and ultima_fecha:
                    self._enviar_mensaje(
                        FaseProceso.FILTRADO,
                        NivelMensaje.INFO,
                        f"📅 Rango real en bandeja: {primera_fecha.date()} a {ultima_fecha.date()}"
                    )
                    
                    # Advertir si el rango solicitado está fuera del disponible
                    if fecha_inicio_normalizada.date() < primera_fecha.date():
                        self._enviar_mensaje(
                            FaseProceso.FILTRADO,
                            NivelMensaje.WARNING,
                            f"⚠️ Fecha inicio ({fecha_inicio_normalizada.date()}) es anterior a la primera fecha en bandeja ({primera_fecha.date()})"
                        )
                    
                    if fecha_fin_normalizada.date() > ultima_fecha.date():
                        self._enviar_mensaje(
                            FaseProceso.FILTRADO,
                            NivelMensaje.WARNING,
                            f"⚠️ Fecha fin ({fecha_fin_normalizada.date()}) es posterior a la última fecha en bandeja ({ultima_fecha.date()})"
                        )
            except Exception as e:
                self._enviar_mensaje(
                    FaseProceso.FILTRADO,
                    NivelMensaje.WARNING,
                    f"No se pudo determinar rango de fechas en bandeja: {str(e)}"
                )
        
        correos_filtrados = []
        
        for idx, item in enumerate(items, 1):
            self._verificar_cancelacion()
            self._verificar_pausa()
            
            # Progreso cada 10%
            if idx % max(1, total_items // 10) == 0 or idx == total_items:
                porcentaje = (idx / total_items) * 100
                self._enviar_mensaje(
                    FaseProceso.FILTRADO,
                    NivelMensaje.INFO,
                    f"Filtrando: {idx}/{total_items} ({porcentaje:.1f}%)"
                )
            
            try:
                # Verificar que sea un correo
                if not hasattr(item, 'ReceivedTime'):
                    continue
                
                # ✅ CORRECCIÓN: Normalizar fecha del correo y comparar correctamente
                fecha_correo = item.ReceivedTime.replace(tzinfo=None)
                
                # Comparar con fechas normalizadas (incluye todo el día)
                if not (fecha_inicio_normalizada <= fecha_correo <= fecha_fin_normalizada):
                    continue
                
                # Filtrar por frases (si se especificaron)
                if frases:
                    asunto = item.Subject.lower() if hasattr(item, 'Subject') else ""
                    if not any(frase.lower() in asunto for frase in frases):
                        continue
                
                # Verificar que tenga adjuntos
                if hasattr(item, 'Attachments') and item.Attachments.Count > 0:
                    correos_filtrados.append(item)
                    
            except Exception as e:
                self.logger.warning(f"Error al procesar item {idx}: {e}")
                continue
        
        self.estadisticas.total_correos = len(correos_filtrados)
        
        self._enviar_mensaje(
            FaseProceso.FILTRADO,
            NivelMensaje.SUCCESS,
            f"Filtrado completado: {len(correos_filtrados)} correos con adjuntos encontrados"
        )
        
        return correos_filtrados
    
    def _descargar_adjuntos(self, correos_filtrados: List, carpeta_destino: str):
        """Descarga adjuntos de los correos filtrados"""
        
        self.fase_actual = FaseProceso.DESCARGA
        self._cambiar_estado(EstadoProceso.PROCESANDO)  # ← Agregado
        
        total_correos = len(correos_filtrados)
        
        self._enviar_mensaje(
            FaseProceso.DESCARGA,
            NivelMensaje.INFO,
            f"Iniciando descarga de adjuntos de {total_correos} correos"
        )
        
        for idx, correo in enumerate(correos_filtrados, 1):
            self._verificar_cancelacion()
            self._verificar_pausa()
            
            try:
                self._procesar_correo(correo, carpeta_destino)
                self.estadisticas.correos_procesados += 1
                
            except Exception as e:
                self._enviar_mensaje(
                    FaseProceso.DESCARGA,
                    NivelMensaje.ERROR,
                    f"Error al procesar correo {idx}: {str(e)}"
                )
            
            # Actualizar progreso
            self._actualizar_progreso(idx, total_correos)
            
            # Log cada 10 correos o al final
            if idx % 10 == 0 or idx == total_correos:
                self._enviar_mensaje(
                    FaseProceso.DESCARGA,
                    NivelMensaje.INFO,
                    f"Procesados: {idx}/{total_correos} correos ({(idx/total_correos)*100:.1f}%)"
                )
        
        self._enviar_mensaje(
            FaseProceso.DESCARGA,
            NivelMensaje.SUCCESS,
            f"Descarga completada: {self.estadisticas.adjuntos_descargados} adjuntos descargados"
        )
    
    def _procesar_correo(self, correo, carpeta_destino: str):
        """Procesa un correo individual y descarga sus adjuntos"""
        
        fecha_correo = correo.ReceivedTime.replace(tzinfo=None)
        carpeta = Path(carpeta_destino)
        
        for adjunto in correo.Attachments:
            self._verificar_cancelacion()
            self._verificar_pausa()
            
            try:
                nombre_archivo = adjunto.FileName
                ruta_archivo = carpeta / nombre_archivo
                
                # Manejar duplicados (método heredado de BackendBase)
                ruta_archivo = self._manejar_nombre_duplicado(ruta_archivo)
                
                # Descargar adjunto
                adjunto.SaveAsFile(str(ruta_archivo))
                
                # Obtener tamaño
                tamaño_mb = ruta_archivo.stat().st_size / (1024 * 1024)
                self.estadisticas.tamaño_total_mb += tamaño_mb
                
                # Registrar descarga
                self.estadisticas.adjuntos_descargados += 1
                self.estadisticas.archivos_descargados.append({
                    'nombre': ruta_archivo.name,
                    'fecha_descarga': datetime.now(),
                    'fecha_correo': fecha_correo.strftime("%d/%m/%Y"),
                    'hora_correo': fecha_correo.strftime("%H:%M:%S")
                })
                
                self._enviar_mensaje(
                    FaseProceso.DESCARGA,
                    NivelMensaje.SUCCESS,
                    f"✓ Descargado: {ruta_archivo.name} ({tamaño_mb:.2f} MB)"
                )
                
            except Exception as e:
                self.estadisticas.adjuntos_fallidos += 1
                self._enviar_mensaje(
                    FaseProceso.DESCARGA,
                    NivelMensaje.ERROR,
                    f"✗ Error al descargar {adjunto.FileName}: {str(e)}"
                )
    
    def _generar_excel_listado(self, carpeta_destino: str):
        """Genera archivo Excel con listado de documentos descargados"""
        
        try:
            if not self.estadisticas.archivos_descargados:
                self._enviar_mensaje(
                    FaseProceso.FINALIZACION,
                    NivelMensaje.INFO,
                    "No hay archivos para listar en Excel"
                )
                return
            
            self._enviar_mensaje(
                FaseProceso.FINALIZACION,
                NivelMensaje.INFO,
                "Generando archivo Excel con listado..."
            )
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Lista de Documentos"
            
            # Encabezados
            encabezados = ["Nº", "Nombre del archivo", "Fecha de descarga", "Fecha correo", "Hora correo"]
            ws.append(encabezados)
            
            # Formato de encabezados
            for col in range(1, 6):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Agregar datos
            for idx, archivo in enumerate(self.estadisticas.archivos_descargados, start=1):
                ws.append([
                    idx,
                    archivo['nombre'],
                    archivo['fecha_descarga'].strftime("%d/%m/%Y %H:%M:%S"),
                    archivo['fecha_correo'],
                    archivo['hora_correo']
                ])
            
            # Crear tabla
            tabla_ref = f"A1:E{len(self.estadisticas.archivos_descargados) + 1}"
            tabla = Table(displayName="TablaDocumentos", ref=tabla_ref)
            estilo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tabla.tableStyleInfo = estilo
            ws.add_table(tabla)
            
            # Autoajustar columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column].width = adjusted_width
            
            # Generar nombre de archivo
            ahora = datetime.now()
            fecha_str = ahora.strftime("%d.%m.%Y")
            hora_str = ahora.strftime("%H.%M.%S")
            nombre_excel = f"lista_documentos_fecha({fecha_str})_hora({hora_str}).xlsx"
            
            ruta_excel = Path(carpeta_destino) / nombre_excel
            wb.save(str(ruta_excel))
            
            self._enviar_mensaje(
                FaseProceso.FINALIZACION,
                NivelMensaje.SUCCESS,
                f"Excel generado: {nombre_excel}"
            )
            
        except Exception as e:
            self._enviar_mensaje(
                FaseProceso.FINALIZACION,
                NivelMensaje.WARNING,
                f"Error al generar Excel: {str(e)}"
            )
    
    # ========================================
    # MÉTODO PÚBLICO PRINCIPAL
    # ========================================
    
    def extraer_adjuntos(self, frases: List[str], destino: str, 
                        outlook_folder: str, fecha_inicio: datetime, 
                        fecha_fin: datetime) -> dict:
        """
        Método principal para extraer adjuntos.
        Usa el template method pattern heredado de BackendBase.
        """
        return self.ejecutar(frases, destino, outlook_folder, fecha_inicio, fecha_fin)


# ========================================
# EJEMPLO DE USO
# ========================================
if __name__ == "__main__":
    import time
    
    def callback_mensaje(fase, nivel, texto: str):
        print(f"[{fase.value.upper()}] [{nivel.value.upper()}] {texto}")
    
    def callback_progreso(actual: int, total: int, porcentaje: float):
        print(f"Progreso: {actual}/{total} ({porcentaje:.1f}%)")
    
    def callback_estado(estado):
        print(f"Estado: {estado.value}")
    
    extractor = ExtractorAdjuntosOutlook(
        callback_mensaje=callback_mensaje,
        callback_progreso=callback_progreso,
        callback_estado=callback_estado
    )
    
    print("\n" + "="*50)
    print("EJEMPLO DE USO - Extractor de Adjuntos Outlook")
    print("="*50)
    print("\n⚠️ Modifica estos valores según tus necesidades:\n")
    
    params_ejemplo = {
        'frases': ['is Signed and Filed', 'se ha firmado'],
        'destino': r'C:\Users\usuario\Downloads\Adjuntos',
        'outlook_folder': r'usuario@empresa.com\Bandeja de entrada\Carpeta',
        'fecha_inicio': datetime(2025, 10, 15),
        'fecha_fin': datetime(2025, 10, 22)
    }
    
    print(f"- Frases: {params_ejemplo['frases']}")
    print(f"- Destino: {params_ejemplo['destino']}")
    print(f"- Carpeta Outlook: {params_ejemplo['outlook_folder']}")
    print(f"- Fechas: {params_ejemplo['fecha_inicio'].date()} a {params_ejemplo['fecha_fin'].date()}")
    print("\n" + "="*50)
    
    respuesta = input("\n¿Deseas ejecutar la extracción con estos valores? (s/n): ")
    
    if respuesta.lower() == 's':
        try:
            print("\nIniciando extracción...")
            stats = extractor.extraer_adjuntos(**params_ejemplo)
            
            print("\n" + "="*50)
            print("RESUMEN FINAL")
            print("="*50)
            print(f"Correos encontrados: {stats['total_correos']}")
            print(f"Correos procesados: {stats['correos_procesados']}")
            print(f"Adjuntos descargados: {stats['adjuntos_descargados']}")
            print(f"Adjuntos fallidos: {stats['adjuntos_fallidos']}")
            print(f"Tamaño total: {stats['tamaño_total_mb']:.2f} MB")
            print(f"Tasa de éxito: {stats['tasa_exito']:.1f}%")
            print(f"Tiempo total: {stats['tiempo_total']:.1f} segundos")
            print("="*50)
                
        except KeyboardInterrupt:
            print("\n\nCancelando...")
            extractor.cancelar()
            time.sleep(2)
        except Exception as e:
            print(f"Error: {e}")
    else:
        print("\nEjecución cancelada.")
    
    print("\nPresiona Enter para salir...")
    input()